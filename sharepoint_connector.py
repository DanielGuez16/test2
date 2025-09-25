import io
import os
from typing import Any, Iterator, Tuple
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import requests
from dotenv import load_dotenv
import tempfile
import pandas as pd

load_dotenv()  # Load environment variables from .env file

# Environment variables
SHAREPOINT_CLIENT_ID = os.environ.get("sharepoint-client-id")
SHAREPOINT_CLIENT_SECRET = os.environ.get("sharepoint-client-secret")
SHAREPOINT_TENANT_ID = os.environ.get("sharepoint-tenant-id")
SITE_RELATIVE_PATH = "/sites/IAForFinance"


class SharePointClient:
    """
    SharePointClient provides methods to interact with SharePoint Online via the Microsoft Graph API.
    This client is designed to manage authentication, file operations, and folder management.
    """

    def __init__(self) -> None:
        """ Initializes the SharePoint client. """
        self.client_id = SHAREPOINT_CLIENT_ID
        self.client_secret = SHAREPOINT_CLIENT_SECRET
        self.tenant_id = SHAREPOINT_TENANT_ID
        self.session = requests.Session()
        self.session.proxies = {
            "http": os.environ.get("HTTP_PROXY"),
            "https": os.environ.get("HTTPS_PROXYy"),
        }
        self.access_token = self.get_access_token()
        self.site_id = self.get_site_id(
            site_hostname="groupebpce.sharepoint.com",
            site_relative_path=SITE_RELATIVE_PATH,
        )

    def get_access_token(self) -> str:
        """Obtains an OAuth2 token for Microsoft Graph API."""
        # Set business proxies to get access token
        business_proxies = {
            "http": os.environ.get("BUSINESS_HTTP_PROXY"),
            "https": os.environ.get("BUSINESS_HTTPS_PROXY"),
        }

        data = {
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        }

        response = requests.post(
            f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token",
            data=data,
            proxies=business_proxies,
            verify=False,
        )
        return response.json().get("access_token")

    def check_token_validity(self) -> str:
        """
        Validates the given access token by making a request to the OAuth2 token endpoint.

        Args:
            access_token (str): The access token to be validated.

        Returns:
            str: The access token if valid, otherwise a new access token is obtained.
        """

        # simple “ping” against Graph API to validate the token
        headers = {"Authorization": f"Bearer {self.access_token}"}
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root"
        response = self.session.get(url, headers=headers)

        if response.status_code == 200:
            return self.access_token
        else:
            access_token = self.get_access_token()
            return access_token
        
    def get_site_id(self, site_hostname: str, site_relative_path: str) -> str:
        """Retrieves the SharePoint site ID.

        Args:
            site_hostname (str): SharePoint hostname.
            site_relative_path (str): Relative path of the site.

        Returns:
            str: Site ID.
        """
        full_url = f"https://graph.microsoft.com/v1.0/sites/{site_hostname}:{site_relative_path}"
        response = self.session.get(full_url, headers={"Authorization": f"Bearer {self.access_token}"})
        return response.json().get("id", "")

    def list_folders_in_path(self, path: str) -> list[str]:
        """Lists the folders at a given path.

        Args:
            path (str): Path in the SharePoint document library.

        Returns:
            list[str]: Names of the folders.
        """
        self.access_token = self.check_token_validity()
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{path}:/children"
        headers = {"Authorization": f"Bearer {self.access_token}"}

        response = self.session.get(url, headers=headers)
        response.raise_for_status()

        # Retrieve folders from the response
        items = response.json().get("value", [])
        folders = [item for item in items if item.get("folder") is not None]
        return [folder.get("name") for folder in folders]

    def list_files_in_path(self, path: str) -> list[str]:
        """Lists the files at a given path.

        Args:
            path (str): Path in the SharePoint document library.

        Returns:
            list[str]: Names of the files.
        """
        self.access_token = self.check_token_validity()
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{path}:/children"
        headers = {
            "Authorization": f"Bearer {self.access_token}",
        }

        response = self.session.get(url, headers=headers)
        response.raise_for_status()
        items = response.json().get("value", [])
        files = [item["name"] for item in items if item.get("file")]
        return files

    # @st.cache_data(show_spinner=False)
    def read_binary_file(self, path: str) -> bytes:
        """Reads a binary file from a SharePoint site using Microsoft Graph API.

        Args:
            path (str): The relative path to the file in the SharePoint site.

        Returns:
            bytes: The binary content of the file.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{path}:/content"
        headers = {"Authorization": f"Bearer {self.access_token}"}

        response = self.session.get(url, headers=headers)
        response.raise_for_status()
        return response.content

    def save_binary_in_sharepoint(self, binary_data: bytes, path: str, get_link: bool = False) -> None:
        """Saves a binary file to a specified path in SharePoint.

        Args:
            binary_data (bytes): The binary content of the file to be saved.
            path (str): The path in SharePoint where the file will be saved.

        Raises:
            requests.exceptions.HTTPError: If the HTTP request to save the file fails.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{path}:/content"
        headers = {
            "Authorization": f"Bearer {self.access_token}",
        }
        response = self.session.put(url, headers=headers, data=binary_data)
        response.raise_for_status()
        # st.session_state.logger.info(f"Binary file successfully saved in SharePoint: {path}")

    # @st.cache_data(show_spinner=False)
    def folder_exists_in_sharepoint(self, folder_path: str) -> bool:
        """Checks if a folder exists in SharePoint.

        Args:
            folder_path (str): Full path of the folder.

        Returns:
            bool: True if it exists, False otherwise.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{folder_path}"
        headers = {"Authorization": f"Bearer {self.access_token}"}
        resp = self.session.get(url, headers=headers)
        return resp.status_code == 200

    # @st.cache_data(show_spinner=False)
    def file_exists_in_sharepoint(self, file_path: str) -> bool:
        """Checks if a file exists in SharePoint.

        Args:
            file_path (str): Full path of the file.

        Returns:
            bool: True if it exists, False otherwise.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{file_path}"
        headers = {"Authorization": f"Bearer {self.access_token}"}
        resp = self.session.get(url, headers=headers)
        return resp.status_code == 200

    def delete_folder_in_sharepoint(self, folder_path: str) -> None:
        """Deletes a folder and all its contents in SharePoint.

        Args:
            folder_path (str): Full path of the folder.

        Returns:
            None.
        """
        # Walk through the folder structure and delete files and subfolders
        for current_path, folders, files in self.walk_sharepoint_path(folder_path):
            # Delete all files in the current folder
            for file_name in files:
                file_path = f"{current_path}/{file_name}"
                self.delete_file_in_sharepoint(file_path)

        # Delete all subfolders (in reverse order to ensure proper deletion)
        for current_path, folders, _ in reversed(list(self.walk_sharepoint_path(folder_path))):
            for folder_name in folders:
                subfolder_path = f"{current_path}/{folder_name}"
                url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{subfolder_path}"
                headers = {"Authorization": f"Bearer {self.access_token}"}
                response = self.session.delete(url, headers=headers)
                response.raise_for_status()

        # Finally, delete the root folder itself
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{folder_path}"
        headers = {"Authorization": f"Bearer {self.access_token}"}
        response = self.session.delete(url, headers=headers)
        response.raise_for_status()

    def save_dataframe_in_sharepoint(self, df: pd.DataFrame, path: str, get_link: bool = False) -> str | None:
        """Saves a DataFrame as an Excel file to a specified path in SharePoint.

        Args:
            df (pd.DataFrame): The DataFrame to be saved as an Excel file.
            path (str): The path in SharePoint where the file will be saved.
            get_link (bool): If True, return the web URL of the saved file.

        Returns:
            str: The web URL of the saved file, if get_link is True.

        Raises:
            requests.exceptions.HTTPError: If the HTTP request to save the file fails.
        """
        # Convert the DataFrame to a binary Excel file
        with BytesIO() as output:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            binary_data = output.getvalue()

        # Call the existing method to save the binary data to SharePoint
        return self.save_binary_in_sharepoint(binary_data, path, get_link)
   

    def delete_file_in_sharepoint(self, file_path: str) -> None:
        """Deletes a file in SharePoint.

        Args:
            file_path (str): Full path of the file.

        Returns:
            None.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{file_path}"
        headers = {"Authorization": f"Bearer {self.access_token}"}

        response = self.session.delete(url, headers=headers)
        response.raise_for_status()

    def walk_sharepoint_path(self, path: str) -> Iterator[Tuple[str, list[str], list[str]]]:
        """Iterator over the hierarchy of a SharePoint path.

        Args:
            path (str): Starting path.

        Yields:
            Iterator[Tuple[str, list[str], list[str]]]: Tuple (path, folders, files).
        """
        folders = self.list_folders_in_path(path)
        files = self.list_files_in_path(path)
        yield path, folders, files

        for folder in folders:
            sub_path = f"{path}/{folder}"
            yield from self.walk_sharepoint_path(sub_path)

    def get_file_last_modified_time(self, file_path: str) -> str:
        """Retrieves the last modified date of a file.

        Args:
            file_path (str): Full path of the file.

        Returns:
            str: Date in ISO 8601 format.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{file_path}"
        headers = {"Authorization": f"Bearer {self.access_token}"}

        response = self.session.get(url, headers=headers)
        response.raise_for_status()

        last_modified_time = response.json().get("lastModifiedDateTime", "")
        return last_modified_time




    # @st.cache_data(show_spinner=False)
    def read_binary_file(self, path: str) -> bytes:
        """Reads a binary file from a SharePoint site using Microsoft Graph API.

        Args:
            path (str): The relative path to the file in the SharePoint site.

        Returns:
            bytes: The binary content of the file.
        """
        self.access_token = self.check_token_validity()
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{path}:/content"
        headers = {"Authorization": f"Bearer {self.access_token}"}

        response = self.session.get(url, headers=headers)
        response.raise_for_status()
        return response.content

    def read_excel_file_as_dict(self, binary_content):
        # Écrire le contenu binaire dans un fichier temporaire
        with open("temp.xlsx", "wb") as f:
            f.write(binary_content)

        # Charger le fichier Excel
        workbook = openpyxl.load_workbook("temp.xlsx")
        
        all_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            data = []
        
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                data.append(row_dict)

            all_data[sheet_name] = data
            
        return all_data
    
    def read_csv_file_as_dict(self, binary_content, encoding='utf-8', delimiter=None, quotechar='"'):
        """
        Lit un fichier CSV à partir du contenu binaire et retourne une liste de dictionnaires.
        
        Args:
            binary_content (bytes): Le contenu binaire du fichier CSV
            encoding (str): L'encodage du fichier (par défaut 'utf-8')
            delimiter (str): Le délimiteur à utiliser (None pour auto-détection)
            quotechar (str): Le caractère de guillemet (par défaut '"')
        
        Returns:
            list: Liste de dictionnaires représentant les données du CSV
        """
        import csv
        import io
        import chardet
        from typing import List, Dict, Any
        
        # Détection automatique de l'encodage si nécessaire
        if encoding == 'auto':
            detected = chardet.detect(binary_content)
            encoding = detected.get('encoding', 'utf-8')
        
        # Tentatives d'encodage multiples pour plus de robustesse
        encodings_to_try = [encoding, 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for enc in encodings_to_try:
            try:
                # Conversion du contenu binaire en string
                text_content = binary_content.decode(enc)
                break
            except UnicodeDecodeError:
                if enc == encodings_to_try[-1]:
                    raise ValueError(f"Impossible de décoder le fichier avec les encodages testés: {encodings_to_try}")
                continue
        
        # Création d'un objet StringIO pour simuler un fichier
        csv_file = io.StringIO(text_content)
        
        # Auto-détection du délimiteur si non spécifié
        if delimiter is None:
            # Lire les premières lignes pour détecter le délimiteur
            sample = csv_file.read(1024)
            csv_file.seek(0)
            
            sniffer = csv.Sniffer()
            try:
                delimiter = sniffer.sniff(sample, delimiters=',;\t|').delimiter
            except csv.Error:
                # Fallback: compter les occurrences de chaque délimiteur possible
                delimiters = [',', ';', '\t', '|']
                delimiter_counts = {d: sample.count(d) for d in delimiters}
                delimiter = max(delimiter_counts.items(), key=lambda x: x[1])[0]
                if delimiter_counts[delimiter] == 0:
                    delimiter = ','  # Délimiteur par défaut
        
        # Lecture du CSV avec gestion d'erreurs
        data = []
        try:
            # Dialecte CSV personnalisé pour plus de robustesse
            csv.register_dialect('robust', 
                            delimiter=delimiter, 
                            quotechar=quotechar,
                            doublequote=True,
                            skipinitialspace=True,
                            quoting=csv.QUOTE_MINIMAL)
            
            reader = csv.DictReader(csv_file, dialect='robust')
            
            # Nettoyage des headers (suppression des espaces)
            if reader.fieldnames:
                reader.fieldnames = [header.strip() if header else f'col_{i}' 
                                for i, header in enumerate(reader.fieldnames)]
            
            for row_num, row in enumerate(reader, start=2):  # Start=2 car ligne 1 = headers
                try:
                    # Nettoyage des valeurs (suppression des espaces)
                    cleaned_row = {k.strip() if k else f'col_{i}': (v.strip() if isinstance(v, str) else v) 
                                for i, (k, v) in enumerate(row.items())}
                    
                    # Conversion automatique des types numériques
                    for key, value in cleaned_row.items():
                        if isinstance(value, str) and value:
                            # Tentative de conversion en nombre
                            try:
                                # Gestion des nombres avec virgule française
                                if ',' in value and '.' not in value:
                                    value = value.replace(',', '.')
                                
                                if '.' in value:
                                    cleaned_row[key] = float(value)
                                else:
                                    cleaned_row[key] = int(value)
                            except ValueError:
                                # Garder comme string si conversion impossible
                                pass
                    
                    data.append(cleaned_row)
                    
                except Exception as e:
                    print(f"Attention: Erreur lors de la lecture de la ligne {row_num}: {e}")
                    # Continuer avec la ligne suivante
                    continue
                    
        except Exception as e:
            raise ValueError(f"Erreur lors de la lecture du CSV: {e}")
        
        finally:
            csv_file.close()
        
        return data

    def read_docx_file_as_text(self, binary_content: bytes) -> str:
        """
        Lit un fichier DOCX depuis du contenu binaire et retourne le texte
        
        Args:
            binary_content: Contenu binaire du fichier DOCX
            
        Returns:
            str: Texte extrait du document
        """
        try:
            from docx import Document
            import io
            
            # Créer un objet Document depuis les bytes
            doc = Document(io.BytesIO(binary_content))
            
            # Extraire tout le texte des paragraphes
            paragraphs = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():  # Ignorer les paragraphes vides
                    paragraphs.append(paragraph.text.strip())
            
            # Joindre tous les paragraphes
            full_text = '\n\n'.join(paragraphs)
            
            return full_text
            
        except Exception as e:
            raise Exception(f"Erreur lecture fichier DOCX: {str(e)}")
        
    def get_file_preview_html(self, path: str) -> dict:
        """Obtient une prévisualisation HTML d'un fichier Office depuis SharePoint
        
        Args:
            path (str): Chemin vers le fichier dans SharePoint
            
        Returns:
            dict: Contient 'html' avec le rendu HTML ou 'error'
        """
        try:
            self.access_token = self.check_token_validity()
            
            # Étape 1: Obtenir les métadonnées du fichier pour récupérer l'ID
            url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{path}"
            headers = {"Authorization": f"Bearer {self.access_token}"}
            
            response = self.session.get(url, headers=headers)
            response.raise_for_status()
            file_data = response.json()
            file_id = file_data.get('id')
            
            if not file_id:
                return {'error': 'File ID not found'}
            
            # Étape 2: Demander la prévisualisation
            preview_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/items/{file_id}/preview"
            
            response = self.session.post(preview_url, headers=headers, json={
                "viewer": "onedrive",  # ou "office" pour une vue plus riche
                "chromeless": True,    # Sans interface OneDrive
                "allowEdit": False     # Lecture seule
            })
            
            if response.status_code == 200:
                preview_data = response.json()
                return {
                    'html': preview_data.get('getUrl', ''),  # URL iframe
                    'embed_url': preview_data.get('postUrl', ''),
                    'success': True
                }
            else:
                return {'error': f'Preview API returned {response.status_code}'}
                
        except Exception as e:
            return {'error': str(e)}


if __name__ == "__main__":
    client = SharePointClient()
    
    file_path = "ALM_Metrics/sources/D_PA_202509110150.csv"
    # file_path = "ALM_Metrics/logs/activity_logs.xlsx"
    data = client.list_files_in_path("ALM_Metrics/sources")
    print(data)

    # try:
    #     # Lire le fichier binaire
    #     binary_content = client.read_binary_file(file_path)

    #     # Lire le contenu du fichier Excel
    #     excel_data = client.read_csv_file_as_dict(binary_content)


    #     # Afficher les données lues
    #     df = pd.DataFrame(excel_data)
    #     print(df)
    #     print(df.to_dict(orient='records'))
    
    # except Exception as e:
    #     print(f"An error occurred: {e}")
